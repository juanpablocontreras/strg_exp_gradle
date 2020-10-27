#save speed data to excel

import os
from openpyxl import Workbook


#Experiment variables
num_str_identifier = "IO"

#functions
def get_exp_max_size(file):
    try:
        num_str = file.split(num_str_identifier, 1)[1]
        return int(num_str)
    except: 
        print("ERROR ON: " + file)
  

#PATHS
logs_path = "/Users/juancontreras/Documents/school/Majumdar_work/size_exp/io_speed"
save_to_path = "/Users/juancontreras/Documents/school/Majumdar_work/size_exp/io_speed.xlsx"

# create workbook and worksheet
wb = Workbook()
ws = wb.active

for (root, dirs, files) in os.walk(logs_path, topdown=True):
    #only one level, where all the files are 
    row = 1
    size_col = 1 #x axis are the dt sizes
    speed_col = 2 #y axis are the speeds corresponding to the size

    for f in files:

        try:
            # get dt max size
            exp_size = get_exp_max_size(f)

            # write dt max size
            ws.cell(row=row, column=size_col, value=int(exp_size))

            # read the file to get average transfer speed
            curr_f = open(logs_path + "/" + f)
            speed = curr_f.read()

            # write speed
            ws.cell(row=row, column=speed_col, value=float(speed))

            row += 1

        except:
            print(f)

        
wb.save(save_to_path)